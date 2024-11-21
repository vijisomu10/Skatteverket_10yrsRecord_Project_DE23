from openpyxl import Workbook, load_workbook
import requests
import os
from reference_extra.api_retreive import get_skatteverket_data

def main_meny():
    get_skatteverket_data()

path = "test_skatteverket.xlsx"
if os.path.exists(path): 
    pass
else:     
    wb = Workbook()
    ws = wb.active
    ws.title = "Skatteverket"
    ws['A1'] = 'Kommun'  # Skapa en rubrik för kommuner i Excel
    """ ws['B1'] = 2014
    ws['C1'] = 2015
    ws['D1'] = 2016
    ws['E1'] = 2017
    ws['F1'] = 2018
    ws['G1'] = 2019
    ws['H1'] = 2020
    ws['I1'] = 2021
    ws['J1'] = 2022
    ws['K1'] = 2023   """
    wb.save(path)  
    wb.close

if __name__ == '__main__':    
    main_meny()
    