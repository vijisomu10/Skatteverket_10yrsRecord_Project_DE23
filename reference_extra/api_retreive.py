from openpyxl import Workbook, load_workbook
import requests
import os

# Funktion för att hämta data från URL
def get_skatteverket_data():
    stockholm_lan_kommuner = []
    lista_med_år = []
    try:    
        offset  = 0
        limit   = 10
        while True:
            url =f'https://skatteverket.entryscape.net/rowstore/dataset/c67b320b-ffee-4876-b073-dd9236cd2a99/json?_offset={offset}&_limit={limit}'
            skatt_data = requests.get(url)
            skatt_data_json = skatt_data.json()             

            if len(skatt_data_json['results'])==0:    
                break
                
            for entry in skatt_data_json['results']:
                kommun = entry['kommun']
                kommun_år = entry['år']
                kommun_skatt = entry['kommunal-skatt']
                stockholm_lan_kommuner.append([kommun, kommun_skatt])
            
                lan_kod = entry['församlings-kod']
                if lan_kod.startswith("01"):
                    stockholm_lan_kommuner.append(kommun)  # Lägg till kommunen i uppsättningen med unika kommuner i Stockholm
                if int(kommun_år) not in lista_med_år: #Om året inte är med i listan
                    lista_med_år.append(int(kommun_år)) #Lägg till det unika året i en lista
            
                for column, value in enumerate(lista_med_år, start = 2):  #Gör så koden skrivs ut på rad 1, frånochmed kolumn 2
                    ws.cell(row=1, column=column, value=value)

            #offset=offset+100 
        return 'passed here'        
    except Exception as err:
        print(err)     

    wb = load_workbook("test_skatteverket.xlsx")
    ws = wb.active    

    print("Stockholms länkommuner:")
    for kommun in stockholm_lan_kommuner:
        print(kommun)
    for kommun_år in lista_med_år:
        print(kommun_år)     
    
    wb.save("test_skatteverket.xlsx")
    wb.close()



    