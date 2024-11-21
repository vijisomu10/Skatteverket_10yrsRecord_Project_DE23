import requests
import json
from openpyxl import Workbook
år_2014=[2014]
år_2015=[2015]
år_2016=[2016]
år_2017=[2017]
år_2018=[2018]
år_2019=[2019]
år_2020=[2020]
år_2021=[2021]
år_2022=[2022]
år_2023=[2023]

wb = Workbook()
ws = wb.active
ws['A1'] = 'Kommuner:'

skatte_info = {}  
def fetch_skatte_data(url, kommuner):
    for year in range(2014, 2024):
        skatte_info[year] = {} 
        for kommun in kommuner:
            skatte_info[year][kommun] = None  

    while url:
        response = requests.get(url)
        data = json.loads(response.text)
        if 'next' in data:
            url = data['next']
        else:
            url = None

        for resultat in data['results']:
            year = int(resultat['år'])
            kommun = resultat['kommun']
            kommunal_skatt = resultat['kommunal-skatt']

            if year >= 2014 and year <= 2023 and kommun in kommuner:
                skatte_info[year][kommun] = kommunal_skatt  

    return skatte_info  

def print_skatte_data(skatte_info):  
    for year in range(2014, 2024):
        print(f"Skattedata för år {year}:")
        for kommun, skatt in skatte_info[year].items():  
            print(f"{kommun}: {skatt}")
            if year == 2014:
                år_2014.append(float(skatt))
            elif year == 2015:
                år_2015.append(float(skatt))
            elif year == 2016:
                år_2016.append(float(skatt))
            elif year == 2017:
                år_2017.append(float(skatt))
            elif year == 2018:
                år_2018.append(float(skatt))
            elif year == 2019:
                år_2019.append(float(skatt))
            elif year == 2020:
                år_2020.append(float(skatt))
            elif year == 2021:
                år_2021.append(float(skatt))
            elif year == 2022:
                år_2022.append(float(skatt))
            elif year == 2023:
                år_2023.append(float(skatt))

def get_skatte_data_for_kommun(skatte_info, years, kommun):         
        if years in skatte_info and kommun in skatte_info[years]:
            skatt = skatte_info[years][kommun]  
            print(f"År {years}, Kommun: {kommun}, Skatt: {skatt}")
        else:
            print(f"Ingen data hittades för år {years} och kommun {kommun}")


url = 'https://skatteverket.entryscape.net/rowstore/dataset/c67b320b-ffee-4876-b073-dd9236cd2a99'
kommuner = ["UPPLANDS VÄSBY", "VALLENTUNA", "ÖSTERÅKER", "VÄRMDÖ", "JÄRFÄLLA", 
            "EKERÖ", "HUDDINGE", "BOTKYRKA", "SALEM", "HANINGE", "TYRESÖ", 
            "UPPLANDS-BRO", "NYKVARN", "TÄBY", "DANDERYD", "SOLLENTUNA", 
            "STOCKHOLM", "SÖDERTÄLJE", "NACKA", "SUNDBYBERG", "SOLNA", 
            "LIDINGÖ", "VAXHOLM", "NORRTÄLJE", "SIGTUNA", "NYNÄSHAMN"]
for kommun in kommuner: # Lägg till alla kommuner i Excel
    ws.append([kommun])



years = 2017

skatte_info = fetch_skatte_data(url, kommuner)
get_skatte_data_for_kommun(skatte_info, years, kommun)
print_skatte_data(skatte_info)

rad=1
for column, value in enumerate(år_2014, start=2):  #Gör så koden skrivs ut på rad 1, frånochmed kolumn 2
    ws.cell(row=rad, column=2, value=value)
    rad+=1
rad=1
for column, value in enumerate(år_2015, start=2):  #Gör så koden skrivs ut på rad 1, frånochmed kolumn 2
    ws.cell(row=rad, column=3, value=value)
    rad+=1
rad=1
for column, value in enumerate(år_2016, start=2):  #Gör så koden skrivs ut på rad 1, frånochmed kolumn 2
    ws.cell(row=rad, column=4, value=value)
    rad+=1
rad=1
for column, value in enumerate(år_2017, start=2):  #Gör så koden skrivs ut på rad 1, frånochmed kolumn 2
    ws.cell(row=rad, column=5, value=value)
    rad+=1
rad=1
for column, value in enumerate(år_2018, start=2):  #Gör så koden skrivs ut på rad 1, frånochmed kolumn 2
    ws.cell(row=rad, column=6, value=value)
    rad+=1
rad=1
for column, value in enumerate(år_2019, start=2):  #Gör så koden skrivs ut på rad 1, frånochmed kolumn 2
    ws.cell(row=rad, column=7, value=value)
    rad+=1
rad=1
for column, value in enumerate(år_2020, start=2):  #Gör så koden skrivs ut på rad 1, frånochmed kolumn 2
    ws.cell(row=rad, column=8, value=value)
    rad+=1
rad=1
for column, value in enumerate(år_2021, start=2):  #Gör så koden skrivs ut på rad 1, frånochmed kolumn 2
    ws.cell(row=rad, column=9, value=value)
    rad+=1
rad=1
for column, value in enumerate(år_2022, start=2):  #Gör så koden skrivs ut på rad 1, frånochmed kolumn 2
    ws.cell(row=rad, column=10, value=value)
    rad+=1
rad=1
for column, value in enumerate(år_2023, start=2):  #Gör så koden skrivs ut på rad 1, frånochmed kolumn 2
    ws.cell(row=rad, column=11, value=value)
    rad+=1

ws.column_dimensions["A"].width = 17
wb.save("skatterna.xlsx")